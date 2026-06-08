"""
update_excel_and_email.py
─────────────────────────
Called by UiPath Studio via "Invoke Python Script" activity.
1. Reads trigger_email.json written by Flask
2. Updates inventory.xlsx with the new/removed item
3. Sends a Gmail email with the full inventory summary
4. Deletes trigger_email.json when done

Usage (UiPath calls this automatically):
    python update_excel_and_email.py
"""

import json
import os
import smtplib
import sys
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

try:
  import openpyxl
except ImportError:
  print("Missing dependency: openpyxl is not installed. Install it with 'pip install openpyxl'.")
  sys.exit(1)
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ══════════════════════════════════════════════════════════════
#  ✏️  CONFIGURE THESE
# ══════════════════════════════════════════════════════════════
GMAIL_USER     = "avishka.inspiredlk@gmail.com"
GMAIL_PASSWORD = "rbxo uhzm hwif mvex"
NOTIFY_TO      = "avishka.inspiredlk@gmail.com"

TRIGGER_FILE   = r"C:\Users\VICTUS\Desktop\inventory-app\backend\trigger_email.json"
EXCEL_FILE     = r"C:\Users\VICTUS\Desktop\inventory-app\inventory.xlsx"


def load_trigger():
    with open(TRIGGER_FILE, "r") as f:
        return json.load(f)


def update_excel(data):
    """Add or remove a row in inventory.xlsx based on the trigger action."""
    action   = data["action"]
    item     = data["item"]
    all_items = data["all_items"]

    # Load or create workbook
    if os.path.exists(EXCEL_FILE):
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inventory"

    # ── Style helpers ─────────────────────────────────────────
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    total_fill  = PatternFill("solid", fgColor="D9EAD3")
    total_font  = Font(bold=True, name="Calibri", size=11)
    center      = Alignment(horizontal="center", vertical="center")
    left        = Alignment(horizontal="left",   vertical="center")
    thin        = Side(style="thin", color="CCCCCC")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Clear sheet and rewrite everything fresh ─────────────
    ws.delete_rows(1, ws.max_row + 1)

    # Headers
    headers = ["Name", "Category", "Quantity", "Price ($)", "Description", "Subtotal ($)"]
    ws.append(headers)
    for col, cell in enumerate(ws[1], 1):
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = center
        cell.border    = border

    # Column widths
    widths = [22, 16, 12, 13, 28, 15]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # Row height
    ws.row_dimensions[1].height = 22

    # Data rows
    alt_fill = PatternFill("solid", fgColor="F8FAFC")
    for row_idx, p in enumerate(all_items, 2):
        row_num = row_idx
        subtotal_formula = f"=C{row_num}*D{row_num}"
        ws.append([
            p["name"],
            p["category"],
            p["quantity"],
            p["price"],
            p.get("description", ""),
            subtotal_formula
        ])
        row = ws[row_idx]
        fill = alt_fill if row_idx % 2 == 0 else None
        for col_idx, cell in enumerate(row, 1):
            cell.border    = border
            cell.alignment = center if col_idx in [3, 4, 6] else left
            if fill:
                cell.fill = fill
            if col_idx in [4, 6]:
                cell.number_format = '"$"#,##0.00'

        # Highlight the changed item
        if p["id"] == item.get("id"):
            hl_color = "D4EDDA" if action == "ADDED" else ("FFE0E0" if action == "REMOVED" else "FFF3CD")
            for cell in row:
                cell.fill = PatternFill("solid", fgColor=hl_color)

    # TOTAL row
    last_data_row = len(all_items) + 1
    total_row     = last_data_row + 1
    ws.append([
        "TOTAL", "",
        f"=SUM(C2:C{last_data_row})",
        "",
        "",
        f"=SUM(F2:F{last_data_row})"
    ])
    for cell in ws[total_row]:
        cell.fill      = total_fill
        cell.font      = total_font
        cell.alignment = center
        cell.border    = border
        if cell.column in [4, 6]:
            cell.number_format = '"$"#,##0.00'

    # Last updated note
    ws[f"A{total_row + 2}"] = f"Last updated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  |  Action: {action} — {item['name']}"
    ws[f"A{total_row + 2}"].font = Font(italic=True, color="888888", size=9)

    wb.save(EXCEL_FILE)
    print(f"[Excel] ✅ Updated: {EXCEL_FILE}")
    return wb


def build_email_html(data):
    action      = data["action"]
    item        = data["item"]
    all_items   = data["all_items"]
    total_qty   = data["total_qty"]
    total_price = data["total_price"]
    timestamp   = data["timestamp"]

    color = {"ADDED": "#22c55e", "REMOVED": "#ef4444", "UPDATED": "#f59e0b"}.get(action, "#6366f1")
    emoji = {"ADDED": "📦",      "REMOVED": "🗑️",      "UPDATED": "✏️"}.get(action, "🔔")

    rows = ""
    for p in all_items:
        subtotal  = p["quantity"] * p["price"]
        highlight = "background:#1e293b;" if p["id"] == item.get("id") else ""
        qty_color = "#f87171" if p["quantity"] < 5 else "#34d399"
        rows += f"""
        <tr style="{highlight}">
          <td style="padding:10px 14px;border-bottom:1px solid #1e293b;color:#94a3b8;font-size:11px;font-family:monospace">#{p['id']}</td>
          <td style="padding:10px 14px;border-bottom:1px solid #1e293b;font-weight:600;color:#f1f5f9">{p['name']}</td>
          <td style="padding:10px 14px;border-bottom:1px solid #1e293b">
            <span style="background:rgba(99,102,241,0.2);color:#818cf8;padding:2px 10px;border-radius:20px;font-size:11px">{p['category']}</span>
          </td>
          <td style="padding:10px 14px;border-bottom:1px solid #1e293b;text-align:center;font-weight:700;color:{qty_color}">{p['quantity']}</td>
          <td style="padding:10px 14px;border-bottom:1px solid #1e293b;text-align:right;color:#2dd4bf;font-family:monospace">${p['price']:.2f}</td>
          <td style="padding:10px 14px;border-bottom:1px solid #1e293b;text-align:right;font-weight:700;color:#f1f5f9;font-family:monospace">${subtotal:.2f}</td>
        </tr>"""

    return f"""<!DOCTYPE html>
<html><head><meta charset="UTF-8"></head>
<body style="margin:0;padding:0;background:#0b1120;font-family:'Segoe UI',Arial,sans-serif">
<div style="max-width:680px;margin:28px auto;border-radius:14px;overflow:hidden;box-shadow:0 8px 40px rgba(0,0,0,0.5)">

  <!-- Header -->
  <div style="background:{color};padding:24px 28px;display:flex;align-items:center;gap:14px">
    <div style="font-size:32px">{emoji}</div>
    <div>
      <div style="font-size:20px;font-weight:800;color:#fff">Inventory {action}</div>
      <div style="font-size:12px;color:rgba(255,255,255,0.75);margin-top:3px">{timestamp}</div>
    </div>
    <div style="margin-left:auto;background:rgba(0,0,0,0.2);border-radius:8px;padding:6px 14px;font-size:12px;color:#fff">
      📊 Excel Updated
    </div>
  </div>

  <!-- Changed item -->
  <div style="background:#0f172a;padding:18px 28px;border-left:4px solid {color}">
    <div style="font-size:11px;color:#64748b;text-transform:uppercase;letter-spacing:.8px;margin-bottom:6px">Item {action.lower()}</div>
    <div style="font-size:20px;font-weight:800;color:#f1f5f9;margin-bottom:6px">{item['name']}</div>
    <div style="display:flex;gap:18px;flex-wrap:wrap">
      <span style="font-size:13px;color:#94a3b8">Category: <strong style="color:#818cf8">{item['category']}</strong></span>
      <span style="font-size:13px;color:#94a3b8">Qty: <strong style="color:#f1f5f9">{item['quantity']}</strong></span>
      <span style="font-size:13px;color:#94a3b8">Price: <strong style="color:#2dd4bf">${item['price']:.2f}</strong></span>
    </div>
  </div>

  <!-- Table -->
  <div style="background:#0f172a;padding:20px 28px">
    <div style="font-size:11px;font-weight:700;color:#475569;text-transform:uppercase;letter-spacing:.8px;margin-bottom:12px">Current Inventory</div>
    <table style="width:100%;border-collapse:collapse;font-size:13px">
      <thead>
        <tr style="background:#1e293b">
          <th style="padding:10px 14px;text-align:left;color:#475569;font-size:10px;text-transform:uppercase;letter-spacing:.8px">ID</th>
          <th style="padding:10px 14px;text-align:left;color:#475569;font-size:10px;text-transform:uppercase;letter-spacing:.8px">Product</th>
          <th style="padding:10px 14px;text-align:left;color:#475569;font-size:10px;text-transform:uppercase;letter-spacing:.8px">Category</th>
          <th style="padding:10px 14px;text-align:center;color:#475569;font-size:10px;text-transform:uppercase;letter-spacing:.8px">Qty</th>
          <th style="padding:10px 14px;text-align:right;color:#475569;font-size:10px;text-transform:uppercase;letter-spacing:.8px">Price</th>
          <th style="padding:10px 14px;text-align:right;color:#475569;font-size:10px;text-transform:uppercase;letter-spacing:.8px">Subtotal</th>
        </tr>
      </thead>
      <tbody>{rows}</tbody>
    </table>
  </div>

  <!-- Totals -->
  <div style="background:#0f172a;padding:0 28px 20px">
    <div style="background:#1e293b;border-radius:12px;padding:16px 20px;display:flex;justify-content:space-around;text-align:center">
      <div>
        <div style="font-size:10px;color:#475569;text-transform:uppercase;letter-spacing:.6px;margin-bottom:4px">Products</div>
        <div style="font-size:26px;font-weight:800;color:#f1f5f9;font-family:monospace">{len(all_items)}</div>
      </div>
      <div style="width:1px;background:#334155"></div>
      <div>
        <div style="font-size:10px;color:#475569;text-transform:uppercase;letter-spacing:.6px;margin-bottom:4px">Total Units</div>
        <div style="font-size:26px;font-weight:800;color:#f1f5f9;font-family:monospace">{total_qty}</div>
      </div>
      <div style="width:1px;background:#334155"></div>
      <div>
        <div style="font-size:10px;color:#475569;text-transform:uppercase;letter-spacing:.6px;margin-bottom:4px">Total Value</div>
        <div style="font-size:26px;font-weight:800;font-family:monospace;color:{color}">${total_price:.2f}</div>
      </div>
    </div>
  </div>

  <!-- Footer -->
  <div style="background:#0b1120;padding:14px 28px;text-align:center;font-size:11px;color:#334155;border-top:1px solid #1e293b">
    Inventory System · Excel auto-updated · Automated by UiPath + Flask
  </div>
</div>
</body></html>"""


def send_email(data):
    action      = data["action"]
    item        = data["item"]
    total_price = data["total_price"]

    subject = f"[Inventory] {action}: {item['name']} | Total: ${total_price:.2f} | Excel Updated ✅"
    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"]    = GMAIL_USER
    msg["To"]      = NOTIFY_TO
    msg.attach(MIMEText(build_email_html(data), "html"))

    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
        server.login(GMAIL_USER, GMAIL_PASSWORD)
        server.sendmail(GMAIL_USER, NOTIFY_TO, msg.as_string())

    print(f"[Email] ✅ Sent → {subject}")


def main():
    print(f"[Bot] 🤖 Running at {datetime.now().strftime('%H:%M:%S')}")

    if not os.path.exists(TRIGGER_FILE):
        print(f"[Bot] ⏭️  No trigger file found. Nothing to do.")
        sys.exit(0)

    try:
        data = load_trigger()
        print(f"[Bot] 📄 Trigger: {data['action']} — {data['item']['name']}")

        update_excel(data)
        send_email(data)

        os.remove(TRIGGER_FILE)
        print(f"[Bot] 🗑️  Trigger file deleted.")
        print(f"[Bot] ✅ Done!")

    except Exception as e:
        print(f"[Bot] ❌ Error: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()