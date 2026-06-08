from flask import Flask, request, jsonify
try:
    from flask_cors import CORS
    CORS_AVAILABLE = True
except Exception:
    CORS_AVAILABLE = False
    def CORS(app, resources=None):
        @app.after_request
        def _cors(response):
            response.headers['Access-Control-Allow-Origin'] = '*'
            response.headers['Access-Control-Allow-Headers'] = 'Content-Type,Authorization'
            response.headers['Access-Control-Allow-Methods'] = 'GET,POST,PUT,DELETE,OPTIONS'
            return response
        return None
import sqlite3, os, json, smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from datetime import datetime

try:
    from google import genai
    from google.genai import types
    GEMINI_AVAILABLE = True
except ImportError:
    GEMINI_AVAILABLE = False
    print("⚠️  google-genai not installed. Run: pip install google-genai")

app = Flask(__name__)
CORS(app)
DB_PATH = os.path.join(os.path.dirname(__file__), "inventory.db")
GEMINI_API_KEY = "AIzaSyB44AS5MXOd2in2LaMbx0NIttiD0TvNsbI"

GMAIL_USER     = "avishka.inspiredlk@gmail.com"
GMAIL_PASSWORD = "rbxo uhzm hwif mvex"
NOTIFY_TO      = "avishka.inspiredlk@gmail.com"
EXCEL_FILE     = r"C:\Users\VICTUS\Desktop\inventory-app\inventory.xlsx"
TRIGGER_FILE   = os.path.join(os.path.dirname(__file__), "trigger_email.json")

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    with get_db() as conn:
        conn.execute("""CREATE TABLE IF NOT EXISTS products (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            name        TEXT    NOT NULL,
            category    TEXT    NOT NULL,
            quantity    INTEGER NOT NULL DEFAULT 0,
            price       REAL    NOT NULL DEFAULT 0.0,
            description TEXT)""")
        conn.execute("""CREATE TABLE IF NOT EXISTS knowledge (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            topic      TEXT    NOT NULL UNIQUE,
            note       TEXT    NOT NULL,
            updated_at DATETIME DEFAULT CURRENT_TIMESTAMP)""")
        conn.commit()

def row_to_dict(row):
    return dict(row)

def update_excel(all_items):
    """Write all inventory items directly to Excel file."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inventory"

        # Header style
        header_fill = PatternFill("solid", fgColor="1E3A5F")
        header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
        center = Alignment(horizontal="center", vertical="center")
        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Headers
        headers = ["Name", "Category", "Quantity", "Price ($)", "Description"]
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border

        # Column widths
        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 13
        ws.column_dimensions["E"].width = 28

        # Data rows
        total_qty = 0
        total_price = 0
        alt_fill = PatternFill("solid", fgColor="EEF2FF")
        for i, item in enumerate(all_items, 2):
            ws.append([
                item["name"],
                item["category"],
                item["quantity"],
                item["price"],
                item.get("description", "")
            ])
            total_qty += item["quantity"]
            total_price += item["quantity"] * item["price"]
            for col, cell in enumerate(ws[i], 1):
                cell.border = border
                cell.alignment = center if col in [3, 4] else Alignment(horizontal="left", vertical="center")
                if i % 2 == 0:
                    cell.fill = alt_fill
                if col == 4:
                    cell.number_format = '"$"#,##0.00'

        # Total row
        total_fill = PatternFill("solid", fgColor="D9EAD3")
        total_font = Font(bold=True, name="Calibri", size=11)
        last = len(all_items) + 2
        ws.append(["TOTAL", "", total_qty, total_price, ""])
        for cell in ws[last]:
            cell.fill = total_fill
            cell.font = total_font
            cell.alignment = center
            cell.border = border
            if cell.column == 4:
                cell.number_format = '"$"#,##0.00'

        # Timestamp
        ws[f"A{last+2}"] = f"Last updated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws[f"A{last+2}"].font = Font(italic=True, color="888888", size=9)

        wb.save(EXCEL_FILE)
        print(f"[Excel] ✅ Updated: {EXCEL_FILE}")

    except Exception as e:
        print(f"[Excel] ❌ Error: {e}")

def write_trigger_file(action, item):
    """Write trigger_email.json so UiPath robot sends email."""
    try:
        with get_db() as conn:
            rows = conn.execute("SELECT * FROM products ORDER BY name").fetchall()
        all_items   = [row_to_dict(r) for r in rows]
        total_qty   = sum(i["quantity"] for i in all_items)
        total_price = sum(i["quantity"] * i["price"] for i in all_items)

        payload = {
            "action"     : action,
            "item"       : item,
            "all_items"  : all_items,
            "total_qty"  : total_qty,
            "total_price": total_price,
            "timestamp"  : datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }

        with open(TRIGGER_FILE, "w") as f:
            json.dump(payload, f, indent=2)

        # Also update Excel directly
        update_excel(all_items)

        print(f"📁 Trigger file written → UiPath will now send email")

    except Exception as e:
        print(f"⚠️  Could not write trigger file: {e}")

def send_inventory_email(action, item_name, quantity, price):
    try:
        with get_db() as conn:
            rows = conn.execute("SELECT * FROM products ORDER BY name").fetchall()
        all_items   = [row_to_dict(r) for r in rows]
        total_qty   = sum(i["quantity"] for i in all_items)
        total_value = sum(i["quantity"] * i["price"] for i in all_items)

        action_color = {"ADDED": "#22c55e", "REMOVED": "#ef4444", "UPDATED": "#f59e0b"}.get(action, "#38bdf8")
        action_icon  = {"ADDED": "➕", "REMOVED": "🗑️", "UPDATED": "✏️"}.get(action, "📦")

        rows_html = ""
        for item in all_items:
            rows_html += f"""
            <tr style="border-bottom:1px solid #1e293b;">
              <td style="padding:8px 12px;">{item['name']}</td>
              <td style="padding:8px 12px;">{item['category']}</td>
              <td style="padding:8px 12px;text-align:center;">{item['quantity']}</td>
              <td style="padding:8px 12px;text-align:right;">${item['price']:.2f}</td>
              <td style="padding:8px 12px;text-align:right;">${item['quantity']*item['price']:.2f}</td>
            </tr>"""

        body = f"""
        <html><body style="font-family:Arial,sans-serif;background:#0f172a;color:#e2e8f0;padding:30px;margin:0;">
        <div style="max-width:650px;margin:auto;background:#1e293b;border-radius:14px;overflow:hidden;">
          <div style="background:#0f172a;padding:24px 30px;border-bottom:3px solid {action_color};">
            <h2 style="margin:0;color:{action_color};">{action_icon} Inventory {action.capitalize()}</h2>
            <p style="margin:6px 0 0;color:#94a3b8;font-size:13px;">Automated by UiPath + Flask ✅</p>
          </div>
          <div style="padding:24px 30px;">
            <div style="background:#0f172a;border-radius:10px;padding:16px 20px;margin-bottom:20px;">
              <p style="margin:0;font-size:18px;font-weight:bold;color:{action_color};">{action_icon} {action} — {item_name}</p>
              <p style="margin:6px 0 0;color:#cbd5e1;">Quantity: <b>{quantity}</b> | Price: <b>${price:.2f}</b></p>
            </div>
            <div style="display:flex;gap:12px;margin-bottom:20px;">
              <div style="flex:1;background:#0f172a;border-radius:10px;padding:14px;text-align:center;">
                <p style="margin:0;color:#94a3b8;font-size:11px;">Total Products</p>
                <p style="margin:4px 0 0;font-size:22px;font-weight:bold;color:#38bdf8;">{len(all_items)}</p>
              </div>
              <div style="flex:1;background:#0f172a;border-radius:10px;padding:14px;text-align:center;">
                <p style="margin:0;color:#94a3b8;font-size:11px;">Total Units</p>
                <p style="margin:4px 0 0;font-size:22px;font-weight:bold;color:#a78bfa;">{total_qty}</p>
              </div>
              <div style="flex:1;background:#0f172a;border-radius:10px;padding:14px;text-align:center;">
                <p style="margin:0;color:#94a3b8;font-size:11px;">Total Value</p>
                <p style="margin:4px 0 0;font-size:22px;font-weight:bold;color:#22c55e;">${total_value:.2f}</p>
              </div>
            </div>
            <table style="width:100%;border-collapse:collapse;background:#0f172a;border-radius:10px;font-size:13px;">
              <thead>
                <tr style="background:#1e3a5f;color:#94a3b8;text-transform:uppercase;font-size:11px;">
                  <th style="padding:10px 12px;text-align:left;">Product</th>
                  <th style="padding:10px 12px;text-align:left;">Category</th>
                  <th style="padding:10px 12px;text-align:center;">Qty</th>
                  <th style="padding:10px 12px;text-align:right;">Price</th>
                  <th style="padding:10px 12px;text-align:right;">Value</th>
                </tr>
              </thead>
              <tbody>{rows_html}</tbody>
            </table>
          </div>
          <div style="padding:16px 30px;background:#0f172a;text-align:center;">
            <p style="margin:0;color:#475569;font-size:11px;">Inventory System · UiPath + Flask</p>
          </div>
        </div></body></html>"""

        msg = MIMEMultipart("alternative")
        msg["Subject"] = f"[Inventory] {action_icon} {action}: {item_name} — Total: ${total_value:.2f}"
        msg["From"]    = GMAIL_USER
        msg["To"]      = NOTIFY_TO
        msg.attach(MIMEText(body, "html"))

        with smtplib.SMTP("smtp.gmail.com", 587) as server:
            server.starttls()
            server.login(GMAIL_USER, GMAIL_PASSWORD)
            server.sendmail(GMAIL_USER, NOTIFY_TO, msg.as_string())
        print(f"📧 Email sent: {action} — {item_name}")

    except Exception as e:
        print(f"⚠️  Email failed: {e}")

@app.route("/api/products", methods=["GET"])
def get_products():
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM products ORDER BY id DESC").fetchall()
    return jsonify([row_to_dict(r) for r in rows])

@app.route("/api/products", methods=["POST"])
def create_product():
    data = request.get_json()
    if not data.get("name") or not data.get("category"):
        return jsonify({"error": "Name and category are required."}), 400
    with get_db() as conn:
        cur = conn.execute(
            "INSERT INTO products (name, category, quantity, price, description) VALUES (?,?,?,?,?)",
            (data["name"], data["category"],
             int(data.get("quantity", 0)),
             float(data.get("price", 0.0)),
             data.get("description", "")))
        conn.commit()
        product = row_to_dict(conn.execute("SELECT * FROM products WHERE id=?", (cur.lastrowid,)).fetchone())
    write_trigger_file("ADDED", product)
    return jsonify(product), 201

@app.route("/api/products/<int:pid>", methods=["PUT"])
def update_product(pid):
    data = request.get_json()
    if not data.get("name") or not data.get("category"):
        return jsonify({"error": "Name and category are required."}), 400
    with get_db() as conn:
        conn.execute(
            "UPDATE products SET name=?,category=?,quantity=?,price=?,description=? WHERE id=?",
            (data["name"], data["category"],
             int(data.get("quantity", 0)),
             float(data.get("price", 0.0)),
             data.get("description", ""), pid))
        conn.commit()
        updated = conn.execute("SELECT * FROM products WHERE id=?", (pid,)).fetchone()
    if updated is None:
        return jsonify({"error": "Not found"}), 404
    product = row_to_dict(updated)
    write_trigger_file("UPDATED", product)
    return jsonify(product)

@app.route("/api/products/<int:pid>", methods=["DELETE"])
def delete_product(pid):
    with get_db() as conn:
        row = conn.execute("SELECT * FROM products WHERE id=?", (pid,)).fetchone()
        if not row:
            return jsonify({"error": "Not found"}), 404
        product = row_to_dict(row)
        conn.execute("DELETE FROM products WHERE id=?", (pid,))
        conn.commit()
    write_trigger_file("REMOVED", product)
    return jsonify({"message": "Deleted."})

@app.route("/api/knowledge", methods=["GET"])
def get_knowledge():
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM knowledge ORDER BY updated_at DESC").fetchall()
    return jsonify([row_to_dict(r) for r in rows])

@app.route("/api/knowledge", methods=["POST"])
def save_knowledge():
    data  = request.get_json()
    topic = data.get("topic", "").strip()
    note  = data.get("note",  "").strip()
    if not topic or not note:
        return jsonify({"error": "Topic and note required."}), 400
    with get_db() as conn:
        conn.execute(
            """INSERT INTO knowledge (topic, note) VALUES (?,?)
               ON CONFLICT(topic) DO UPDATE SET
               note=excluded.note, updated_at=CURRENT_TIMESTAMP""",
            (topic.lower(), note))
        conn.commit()
        row = conn.execute("SELECT * FROM knowledge WHERE topic=?", (topic.lower(),)).fetchone()
    return jsonify(row_to_dict(row)), 201

@app.route("/api/knowledge/<int:kid>", methods=["DELETE"])
def delete_knowledge(kid):
    with get_db() as conn:
        conn.execute("DELETE FROM knowledge WHERE id=?", (kid,))
        conn.commit()
    return jsonify({"message": "Deleted."})

@app.route("/api/chat", methods=["POST"])
def chat():
    if not GEMINI_AVAILABLE:
        return jsonify({"error": "Gemini SDK not installed. Run: pip install google-genai"}), 500

    data         = request.get_json()
    user_message = (data.get("message") or "").strip()
    if not user_message:
        return jsonify({"error": "Message is required."}), 400

    try:
        with get_db() as conn:
            product_rows   = conn.execute("SELECT * FROM products ORDER BY category, name").fetchall()
            knowledge_rows = conn.execute("SELECT topic, note FROM knowledge ORDER BY updated_at DESC").fetchall()
    except Exception as db_err:
        return jsonify({"error": f"Database error: {db_err}"}), 500

    products  = [row_to_dict(r) for r in product_rows]
    knowledge = [row_to_dict(r) for r in knowledge_rows]

    if products:
        lines = []
        for p in products:
            line = (f"- {p['name']} | Category: {p['category']} "
                    f"| Qty: {p['quantity']} | Price: ${p['price']:.2f}")
            if p["description"]:
                line += f" | {p['description']}"
            lines.append(line)
        total_value = sum(p["quantity"] * p["price"] for p in products)
        low_stock   = [p["name"] for p in products if p["quantity"] < 6]
        summary = (
            f"Total products: {len(products)}\n"
            f"Total inventory value: ${total_value:.2f}\n"
            f"Low stock (qty < 5): {', '.join(low_stock) if low_stock else 'None'}\n\n"
            f"Products:\n" + "\n".join(lines)
        )
    else:
        summary = "Inventory is empty."

    if knowledge:
        knowledge_text = "OWNER'S PERSONAL NOTES (highest priority):\n"
        for k in knowledge:
            knowledge_text += f"- {k['topic']}: {k['note']}\n"
    else:
        knowledge_text = "OWNER'S PERSONAL NOTES: (none saved yet)"

    system_prompt = f"""You are a helpful inventory assistant for a small business owner.

STRICT RULES:
1. ALWAYS check Owner's Personal Notes FIRST before answering anything.
2. If notes have relevant info → use it and say "Based on your notes: ..."
3. If notes have nothing relevant → use AI knowledge and say "Based on general knowledge: ..."
4. ONLY save knowledge if the owner clearly says: "remember that...", "note that...", "save that...", "don't forget that..."
   If they say these, respond ONLY with:
   SAVE_KNOWLEDGE:topic=<topic>|note=<the fact to save>
5. NEVER save knowledge from casual conversation or questions.

{knowledge_text}

Current inventory:
{summary}"""

    try:
        client   = genai.Client(api_key=GEMINI_API_KEY)
        response = client.models.generate_content(
            model="gemini-2.5-flash",
            config=types.GenerateContentConfig(
                system_instruction=system_prompt,
                max_output_tokens=1000,
            ),
            contents=user_message,
        )
        reply = (response.text or "").strip()
    except Exception as gemini_err:
        return jsonify({"error": f"Gemini error: {gemini_err}"}), 500

    knowledge_updated = False
    if reply.startswith("SAVE_KNOWLEDGE:") and "|note=" in reply:
        try:
            parts      = reply.replace("SAVE_KNOWLEDGE:", "").strip()
            topic_part, note_part = parts.split("|note=", 1)
            topic = topic_part.replace("topic=", "").strip()
            note  = note_part.strip()
            if len(topic) > 2 and len(note) > 5:
                with get_db() as conn:
                    conn.execute(
                        """INSERT INTO knowledge (topic, note) VALUES (?,?)
                           ON CONFLICT(topic) DO UPDATE SET
                           note=excluded.note, updated_at=CURRENT_TIMESTAMP""",
                        (topic.lower(), note))
                    conn.commit()
                reply             = f"✅ Got it! I've saved this to your notes!\n**{topic}**: {note}"
                knowledge_updated = True
            else:
                reply = "I understood, but that was too vague to save. Can you be more specific?"
        except Exception:
            reply = "I understood that, but had trouble saving it. Please try again."

    return jsonify({"reply": reply, "knowledge_updated": knowledge_updated})

if __name__ == "__main__":
    init_db()
    print("✅  Database ready.")
    print("🚀  Flask running at http://localhost:5000")
    print(f"📁  Trigger file location: {TRIGGER_FILE}")
    app.run(debug=True, port=5000)