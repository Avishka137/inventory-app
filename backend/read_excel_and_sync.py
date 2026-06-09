"""
read_excel_and_sync.py
- Reads inventory.xlsx
- Adds new rows (Status empty) to inventory via Flask API
- Removes items from inventory if their row was deleted from Excel
- Marks added rows as DONE
- Sends email summary
"""

import openpyxl
import urllib.request
import urllib.parse
import json
import smtplib
import os
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from datetime import datetime
from openpyxl.styles import PatternFill, Font

EXCEL_FILE  = r"C:\Users\VICTUS\Desktop\inventory-app\inventory.xlsx"
API_BASE    = "http://localhost:5000/api/products"
GMAIL_USER  = "avishka.inspiredlk@gmail.com"
GMAIL_PASS  = "rbxo uhzm hwif mvex"
NOTIFY_TO   = "avishka.inspiredlk@gmail.com"

def api_get():
    with urllib.request.urlopen(API_BASE) as r:
        return json.loads(r.read())

def api_add(name, category, quantity, price, description):
    data = json.dumps({
        "name": name, "category": category,
        "quantity": int(quantity), "price": float(price),
        "description": description
    }).encode()
    req = urllib.request.Request(API_BASE, data=data,
          headers={"Content-Type": "application/json"}, method="POST")
    with urllib.request.urlopen(req) as r:
        return json.loads(r.read())

def api_delete(pid):
    req = urllib.request.Request(f"{API_BASE}/{pid}", method="DELETE")
    urllib.request.urlopen(req)

def send_email(added, removed):
    if not added and not removed:
        return
    body = "<h2>Excel to Inventory Sync Report</h2>"
    if added:
        body += "<h3>Added to Inventory</h3><ul>"
        for i in added:
            body += f"<li><b>{i}</b></li>"
        body += "</ul>"
    if removed:
        body += "<h3>Removed from Inventory</h3><ul>"
        for i in removed:
            body += f"<li><b>{i}</b></li>"
        body += "</ul>"
    body += f"<p style='color:gray'>Synced at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>"

    msg = MIMEMultipart("alternative")
    msg["Subject"] = f"[UiPath] Inventory Sync: +{len(added)} added, -{len(removed)} removed"
    msg["From"]    = GMAIL_USER
    msg["To"]      = NOTIFY_TO
    msg.attach(MIMEText(body, "html"))

    with smtplib.SMTP("smtp.gmail.com", 587) as s:
        s.starttls()
        s.login(GMAIL_USER, GMAIL_PASS)
        s.sendmail(GMAIL_USER, NOTIFY_TO, msg.as_string())
    print(f"[Email] Sent: +{len(added)} added, -{len(removed)} removed")

def read_and_sync():
    print(f"[Bot] Running at {datetime.now().strftime('%H:%M:%S')}")

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    # Ensure Status column header exists
    if ws.cell(1, 6).value != "Status":
        ws.cell(1, 6).value = "Status"

    # Read all rows from Excel (skip header, skip TOTAL/Last updated rows)
    excel_items = {}  # name -> row_number
    for r in range(2, ws.max_row + 1):
        name = str(ws.cell(r, 1).value or "").strip()
        if not name or name.upper() in ("TOTAL", "LAST UPDATED", ""):
            continue
        if name.startswith("Last updated"):
            continue
        excel_items[name.lower()] = {
            "row": r,
            "name": name,
            "category": str(ws.cell(r, 2).value or "General").strip(),
            "quantity": ws.cell(r, 3).value or 1,
            "price":    ws.cell(r, 4).value or 0,
            "description": str(ws.cell(r, 5).value or "").strip(),
            "status": str(ws.cell(r, 6).value or "").strip()
        }

    # Get current inventory from Flask
    try:
        db_items = api_get()
    except Exception as e:
        print(f"[API] Error connecting to Flask: {e}")
        return

    db_names = {p["name"].lower(): p for p in db_items}

    added   = []
    removed = []

    # ADD: rows in Excel with no Status → add to inventory
    done_fill = PatternFill("solid", fgColor="00B050")
    done_font = Font(bold=True, color="FFFFFF")

    for key, item in excel_items.items():
        if item["status"] != "DONE":
            try:
                result = api_add(item["name"], item["category"],
                                 item["quantity"], item["price"],
                                 item["description"])
                r = item["row"]
                ws.cell(r, 6).value = "DONE"
                ws.cell(r, 6).fill  = done_fill
                ws.cell(r, 6).font  = done_font
                added.append(item["name"])
                print(f"[API] Added: {item['name']}")
            except Exception as e:
                print(f"[API] Failed to add {item['name']}: {e}")

    # REMOVE: items in DB that are no longer in Excel → delete from inventory
    for key, product in db_names.items():
        if key not in excel_items:
            try:
                api_delete(product["id"])
                removed.append(product["name"])
                print(f"[API] Removed: {product['name']}")
            except Exception as e:
                print(f"[API] Failed to remove {product['name']}: {e}")

    try:
        wb.save(EXCEL_FILE)
        print("[Excel] Saved")
    except PermissionError:
        print("[Excel] ERROR: Close Excel first, then run again!")
        return

    send_email(added, removed)

    if not added and not removed:
        print("[Bot] No changes detected.")

if __name__ == "__main__":
    read_and_sync()